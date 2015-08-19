#! python3
# excelGetCells.py - example on how to get data/work with data from an excel sheet, run th

import openpyxl
from openpyxl.cell import get_column_letter, column_index_from_string
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print('Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value)
print('Cell ' + c.coordinate + ' is ' + c.value)
print(sheet['C1'].value)
print(sheet.get_highest_row())
print(sheet.get_highest_column())
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(sheet.get_highest_column()))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))

